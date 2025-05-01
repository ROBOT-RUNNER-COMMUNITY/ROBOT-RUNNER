import os
import subprocess
import time
from openpyxl import Workbook
from robot.api import ExecutionResult
from PyQt6.QtCore import Qt, QTimer
from PyQt6 import QtCore
from PyQt6.QtGui import QMovie, QPixmap
from PyQt6.QtWidgets import QListWidgetItem, QMessageBox
from utils.resource_utils import resource_path
from utils.display_utils import show_cross
from openpyxl.chart import BarChart, Reference
import traceback
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def load_tests(window):
    window.testList.clear()
    window.loadingLabel.show()
    loading_gif_path = resource_path("images/loading.gif")
    window.loading_movie = QMovie(loading_gif_path)
    window.loading_movie.setScaledSize(QtCore.QSize(45, 45))
    window.loadingLabel.setMovie(window.loading_movie)
    window.loading_movie.start()
    QTimer.singleShot(1500, lambda: populate_tests(window))

def populate_tests(window):
    window.loading_movie.stop()
    window.loadingLabel.clear()

    if window.test_directory:
        test_files = [file for file in os.listdir(window.test_directory) if file.endswith(".robot")]

        if not test_files:
            show_cross(window)
            window.label.setStyleSheet("color: #ad402a")
        else:
            for file in test_files:
                item = QListWidgetItem(file)
                item.setCheckState(Qt.CheckState.Unchecked)
                window.testList.addItem(item)

            window.label.setStyleSheet("color: green")
            check_icon_path = resource_path("images/check.png")
            check_pixmap = QPixmap(check_icon_path).scaled(27, 27, QtCore.Qt.AspectRatioMode.KeepAspectRatio)
            window.loadingLabel.setPixmap(check_pixmap)
            window.loadingLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)

def run_tests(window):
    try:
        if not window.test_directory:
            window.resultLabel.setStyleSheet("color: none")
            window.resultLabel.setText("Veuillez sélectionner un dossier.")
            window.resultLabel.setStyleSheet("color: #ad402a; font: bold")
            return False
        
        if not window.output_directory:
            window.resultLabel.setStyleSheet("color: none")
            window.resultLabel.setText("Veuillez sélectionner un emplacement pour les résultats.")
            window.resultLabel.setStyleSheet("color: #ad402a; font: bold")
            return False
        
        selected_tests = [os.path.join(window.test_directory, window.testList.item(i).text())
                         for i in range(window.testList.count())
                         if window.testList.item(i).checkState() == Qt.CheckState.Checked]
        
        if not selected_tests:
            window.resultLabel.setStyleSheet("color: none")
            window.resultLabel.setText("Veuillez sélectionner au moins un test.")
            window.resultLabel.setStyleSheet("color: #ad402a; font: bold")
            return False
        
        num_processes = window.processInput.text()
        repport_title = "AUTOS TESTS - REPORT"
        log_title = "AUTOS TESTS - LOG"

        # Ensure output directory exists
        os.makedirs(window.output_directory, exist_ok=True)

        if num_processes == "1":
            command = ["robot", "-d", window.output_directory] + selected_tests
        else:
            command = ["pabot", "--processes", num_processes, "--outputdir", window.output_directory, 
                      "--reporttitle", repport_title, "--logtitle", log_title] + selected_tests
        
        # Run tests
        process = subprocess.Popen(command, cwd=window.test_directory)
        process.wait()

        # Wait for output.xml to be generated
        output_path = os.path.join(window.output_directory, "output.xml")
        max_wait = 15  
        wait_interval = 0.5  

        for _ in range(int(max_wait / wait_interval)):
            if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                break
            time.sleep(wait_interval)
        else:
            raise FileNotFoundError(f"output.xml not generated after {max_wait} seconds")

        # Process results
        result = ExecutionResult(output_path)
        
        if result.suite.statistics.failed >= 1:
            window.resultLabel.setStyleSheet("color: none")
            window.resultLabel.setText(f"Total: {result.suite.statistics.total} | Passés: {result.suite.statistics.passed} | Échoués: {result.suite.statistics.failed}")
            window.resultLabel.setStyleSheet("color: #ad402a; font: bold")
        else:
            window.resultLabel.setStyleSheet("color: none")
            window.resultLabel.setText(f"Total: {result.suite.statistics.total} | Passés: {result.suite.statistics.passed} | Échoués: {result.suite.statistics.failed}")
            window.resultLabel.setStyleSheet("color: green; font: bold")
        
        return True

    except Exception as e:
        QMessageBox.critical(
            window,
            "Test Execution Error",
            f"Failed to run tests: {str(e)}"
        )
        return False

def open_report(window):
    report_path = os.path.join(window.output_directory, "report.html")
    if os.path.exists(report_path):
        os.system(f'start "" "{report_path}"')

def open_log(window):
    log_path = os.path.join(window.output_directory, "log.html")
    if os.path.exists(log_path):
        os.system(f'start "" "{log_path}"')

def export_results(window):
    """Export Robot Framework test results to an Excel file with comprehensive error handling."""
    try:
        # Validate output directory
        if not hasattr(window, 'output_directory') or not window.output_directory:
            window.resultLabel.setText("Output directory not set!")
            window.resultLabel.setStyleSheet("color: #ad402a")
            return

        # Check if output.xml exists
        output_xml = os.path.join(window.output_directory, "output.xml")
        if not os.path.exists(output_xml):
            window.resultLabel.setText("Run tests first to generate output.xml!")
            window.resultLabel.setStyleSheet("color: #ad402a")
            return

        # Create results directory if needed
        try:
            os.makedirs(window.output_directory, exist_ok=True)
        except OSError as e:
            window.resultLabel.setText(f"Cannot create directory: {str(e)}")
            window.resultLabel.setStyleSheet("color: #ad402a")
            return

        # Process test results
        result = ExecutionResult(output_xml)
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Set up headers with formatting
        headers = ["Suite Name", "Test Name", "Status", "Duration (s)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        passed_count = 0
        failed_count = 0

        # Populate test data with status formatting
        for suite in result.suite.suites:
            for test in suite.tests:
                status = "Passed" if test.passed else "Failed"
                duration = getattr(test.elapsed_time, 'total_seconds', lambda: 0)()
                ws.append([suite.name, test.name, status, duration])

                status_cell = ws.cell(row=ws.max_row, column=3)
                if test.passed:
                    passed_count += 1
                    status_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                else:
                    failed_count += 1
                    status_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")

        # Add summary section
        ws.append([])  # Empty row
        summary_headers = ["Status", "Count"]
        ws.append(summary_headers)

        # Format summary headers
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
        
        # Add summary data
        ws.append(["Passed", passed_count])
        ws.append(["Failed", failed_count])
        
        # Format summary rows
        for row in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row):
            for cell in row:
                if cell.column == 1:
                    cell.fill = PatternFill(start_color="E7E6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")

        # Create bar chart
        chart_data_start = ws.max_row + 2
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Test Results Summary"
        chart.y_axis.title = "Number of Tests"
        chart.x_axis.title = "Status"
        chart.grouping = "clustered"
        chart.overlap = 0
        chart.width = 15
        chart.height = 10

        data = Reference(ws,
                        min_col=2,
                        min_row=chart_data_start - 3,
                        max_row=chart_data_start - 2)
        cats = Reference(ws,
                        min_col=1,
                        min_row=chart_data_start - 3,
                        max_row=chart_data_start - 2)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        ws.add_chart(chart, "F2")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        # Define output path
        excel_filename = 'test_results.xlsx'
        excel_path = os.path.join(window.output_directory, excel_filename)
        
        # Attempt to save with error handling
        try:
            wb.save(excel_path)
        except PermissionError:
            window.resultLabel.setText("Please close the Excel file to generate new report")
            window.resultLabel.setStyleSheet("color: #ad402a")
            return
        except Exception as e:
            window.resultLabel.setText("Error saving report file")
            window.resultLabel.setStyleSheet("color: #ad402a")
            return

        # Verify and show success
        if os.path.exists(excel_path):
            window.resultLabel.setText(f"Report saved to {excel_path}")
            window.resultLabel.setStyleSheet("color: green")
            try:
                os.startfile(excel_path)
            except:
                pass  
        else:
            window.resultLabel.setText("Failed to create report file")
            window.resultLabel.setStyleSheet("color: #ad402a")

    except Exception:
        window.resultLabel.setText("Unexpected error during export")
        window.resultLabel.setStyleSheet("color: #ad402a")